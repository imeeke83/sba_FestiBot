# -*- coding: utf-8 -*-
"""
Created on Thu Sep 19 09:15:31 2019

@author: KangmoonKo
"""

import requests
from flask import Flask, request, Response
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta
import calendar
import copy

app_festibot = Flask(__name__)

# Telegram API Key
API_KEY = ''

# API 호출을 위한 변수
serviceKey='' # API Key
numOfRows=20 # 20행 보기
listYN='Y' # 목록으로 보기
arrange='O' # 대표 이미지가 반드시 있는 정렬 (O=제목순, P=조회순)

# 상태정보 저장을 위한 DataFrame
stateDic = {'stateCode':[], 'eventStartDate':[], 'eventEndDate':[], 'contentCode':[]}
stateDB = pd.DataFrame(stateDic, columns=['stateCode', 'eventStartDate', 'eventEndDate', 'contentCode'])
# User DataBase
EXCEL_FILE_NAME = "UserDB.xlsx"
db = load_workbook(filename=EXCEL_FILE_NAME) # 기존에 존재하는 엑셀파일만 불러올 수 있음
userInfoDB = db['userInfoDB'] # db 엑셀파일 안의 userInfoDB 시트
# 데이터 연산을 위한 List
dateList = ['오늘 축제', '내일 축제', '이번주 축제', '이번달 축제']
dateWrite = re.compile('([12]\d{3}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01]))-([12]\d{3}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01]))')
contentListName = ['문화관광', '일반', '전통공연', '연극', '뮤지컬', '오페라', '전시회', '박람회', '컨벤션', '무용', '클래식음악회', '대중콘서트', '영화', '스포츠경기', '기타행사']
contentListCode = ['A02070100', 'A02070200', 'A02080100', 'A02080200', 'A02080300', 'A02080400', 'A02080500', 'A02080600', 'A02080700', 'A02080800', 'A02080900', 'A02081000', 'A02081100', 'A02081200', 'A02081300']

# 사용자에게 메시지를 받아, 필요한 항목을 불러옴
def parse_message(message):
    """
    사용자에게 메시지를 받아, 필요한 항목을 불러옴
    """
    print('parse_message')
    user_id = message['message']['chat']['id']
    userName = message['message']['chat']['first_name']+message['message']['chat']['last_name']
    msg = message['message']['text']
    
    return user_id, userName, msg

# 사용자에게 메세지를 보냄
def send_message(user_id, text):
    """
    사용자에게 메세지를 보냄
    """
    print('send_message')
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
    params = {'chat_id': user_id, 'text': text}

    response = requests.post(url, json=params)
    return response

# UserDB.xlsx 파일에 User 정보가 존재하는지 확인
# 존재하지 않는 경우, User 정보를 추가하고 초기화
def find_userInfo(user_id, userName):
    """
    UserDB.xlsx 파일에 User 정보가 존재하는지 확인
    존재하지 않는 경우, User 정보를 추가하고 초기화
    """
    print('find_userInfo')
    for row in userInfoDB.rows:  
        if row[0].value == user_id: # 방문했던 적이 있었던 사람, 리스트로 저장되므로 리스트로 접근
            return True
    userInfoDB[userInfoDB.max_row+1][0].value = user_id
    userInfoDB[userInfoDB.max_row][1].value = userName
    for i in range(3, 18):
        userInfoDB.cell(row=userInfoDB.max_row, column=i).value = 0
    db.save(EXCEL_FILE_NAME)
    return False

# 저장된 User의 선호 축제를 검색하여 선호 축제의 코드값 검색
def find_whatUserLike(user_id):
    """
    저장된 User의 선호 축제를 검색하여 선호 축제의 코드값 검색
    """
    print('find_whatUserLike')
    for row in userInfoDB.rows:  
        if row[0].value == user_id: # 방문했던 적이 있었던 사람, 리스트로 저장되므로 리스트로 접근
            userRow = row[0].row # 행에 대한 정보를 유저로우에 담아줌(사용자 정보 추가)

    userLikecontent = []
    
    for i in range(3, 18):
        userLikecontent.append(userInfoDB.cell(row=userRow, column=i).value)
        
    max = 0
    for value in userLikecontent:
        if value > max:
            max = value

    if max > 0:
        return contentListCode.index(userLikecontent.index(max))
    else:
        return 0

# 처음 방문한 사용자에게 환영 메시지 출력.
def send_welcome_msg(user_id, userName):
    """
    처음 방문한 사용자에게 환영 메시지 출력.
    """
    print('send_welcome_msg')
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
    welcomeMsg = f'{userName}님 안녕하세요. 저는 페스티봇이에요. 축제를 알려드립니다 !'
    params = {'chat_id': user_id, 'text':welcomeMsg}
    requests.post(url, json=params)

# 유저 정보를 확인하여, 기존 유저 또는 첫 방문자인지 검사
def thisUserIsFirst(user_id, userName):
    """
    유저 정보를 확인하여, 기존 유저 또는 첫 방문자인지 검사
    """
    print('thisUserIsFirst')
    if find_userInfo(user_id, userName):
        userLike = find_whatUserLike(user_id)
        if userLike == 0:
            pass
        else:
            stateDB.loc[user_id, 'contentCode'] = userLike
    else:
        send_welcome_msg(user_id, userName)
    
    if user_id in stateDB.index:
        pass
    else:
        stateDB.loc[user_id] = np.nan

# 사용자에게 최초 버튼 선택 화면을 보여줌
def click_buttonFirst(user_id, msg):
    """
    사용자에게 최초 버튼 선택 화면을 보여줌
    """
    print('click_buttonFirst')
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)   #sendMessage
    keyboard = {                                        # Keyboard 형식
            'keyboard':[[{'text': '축제 기간'}, {'text': '축제 종류'}]],
            'one_time_keyboard' : True
            }
    
    params = {'chat_id':user_id, 'text':msg, 'reply_markup' : keyboard}
    requests.post(url, json=params)


############################## 일정 기준 검색 ##############################

# 사용자에게 세부 일정 검색 선택 화면을 보여줌
def choice_calendarDate(user_id, msg):
    """
    사용자에게 세부 일정 검색 선택 화면을 보여줌
    """
    print('choice_calendarDate')
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)   #sendMessage
    keyboard = {                                        # Keyboard 형식
            'keyboard':[[{'text': '오늘 축제'}, {'text': '내일 축제'}],
                        [{'text': '이번주 축제'}, {'text': '이번달 축제'}]],
            'one_time_keyboard' : True
            }
    
    params = {'chat_id':user_id, 'text': msg, 'reply_markup' : keyboard}
    requests.post(url, json=params)   
    
# 사용자가 선택한 세부 일정 별 시작일 및 종료일을 state로 저장
def choice_fixCalendarDate(user_id, msg):
    """
    사용자가 선택한 세부 일정 별 시작일 및 종료일을 state로 저장
    """
    print('choice_fixCalendarDate')
    if dateWrite.match(msg):
        stateDB.loc[user_id, 'eventStartDate'], stateDB.loc[user_id, 'eventEndDate'] = msg.split('-')
        stateDB.loc[user_id, 'stateCode'] = np.nan
    elif msg == '오늘 축제':
        stateDB.loc[user_id, 'eventStartDate'] = datetime.today().strftime("%Y%m%d")
        stateDB.loc[user_id, 'eventEndDate'] = datetime.today().strftime("%Y%m%d")
        stateDB.loc[user_id, 'stateCode'] = np.nan
    elif msg == '내일 축제':
        tomorrow = datetime.today() + timedelta(days=1)
        stateDB.loc[user_id, 'eventStartDate'] = tomorrow.strftime("%Y%m%d")
        stateDB.loc[user_id, 'eventEndDate'] = tomorrow.strftime("%Y%m%d")
        stateDB.loc[user_id, 'stateCode'] = np.nan
    elif msg == '이번주 축제':
        startDate = datetime.today() - timedelta(days=datetime.today().weekday())
        endDate = datetime.today() - timedelta(days=datetime.today().weekday()-7)
        stateDB.loc[user_id, 'eventStartDate'] = startDate.strftime("%Y%m%d")
        stateDB.loc[user_id, 'eventEndDate'] = endDate.strftime("%Y%m%d")
        stateDB.loc[user_id, 'stateCode'] = np.nan
    elif msg == '이번달 축제':
        startDate = datetime.today().replace(day=1)
        endDate = datetime.today().replace(day=calendar.monthrange(datetime.today().year, datetime.today().month)[1])
        stateDB.loc[user_id, 'eventStartDate'] = startDate.strftime("%Y%m%d")
        stateDB.loc[user_id, 'eventEndDate'] = endDate.strftime("%Y%m%d")
        stateDB.loc[user_id, 'stateCode'] = np.nan

    print(stateDB)


############################## 종류 기준 검색 ##############################
    
# 사용자가 선택한 종류를 state로 저장
def choice_contentCode(user_id, msg):
    """
    사용자가 선택한 종류를 state로 저장
    """
    print('choice_contentCode')
    if msg in contentListName:
        index = contentListName.index(msg)
        stateDB.loc[user_id, 'contentCode'] = contentListCode[index]
        stateDB.loc[user_id, 'stateCode'] = np.nan
    elif msg.isdigit() and int(msg)>0 and int(msg)<16:
        index = int(msg)-1
        stateDB.loc[user_id, 'contentCode'] = contentListCode[index]
        stateDB.loc[user_id, 'stateCode'] = np.nan
        
    print(stateDB)


############################## API 활용 ##############################

# 전체 축제 중 사용자가 선택한 종류의 축제만을 선별
def searchContentFestival(user_id, startDate, endDate, content, pageNo):
    """
    전체 축제 중 사용자가 선택한 종류의 축제만을 선별
    """
    print('searchContentFestival')
    url=f'http://api.visitkorea.or.kr/openapi/service/rest/KorService/searchFestival?numOfRows={numOfRows}&MobileOS=ETC&MobileApp=Festibot&serviceKey={serviceKey}&listYN={listYN}&arrange={arrange}&areaCode=1&eventStartDate={startDate}&eventEndDate={endDate}&pageNo={pageNo}&_type=json'
    resp = requests.get(url)
    data = resp.json()
    festivalInfo = data['response']['body']['items']['item']
    festivalList = []

    for infoDec in festivalInfo:
        if infoDec['cat3'] == content:
            fixList = { 
                'cat3':infoDec['cat3'], 'firstimage':infoDec['firstimage'], 'title':infoDec['title'],
                'eventenddate':infoDec['eventenddate'], 'eventstartdate':infoDec['eventstartdate'],
                'addr1':infoDec['addr1'] } #, 'tel':infoDec['tel'] }
            festivalList.append(copy.deepcopy(fixList))
    print('searchContentFestival 1 : ',festivalList)
    return festivalList

# 조건에 맞는 모든 축제 검색
def searchAllFestival(user_id, startDate, endDate, content):
    """
    조건에 맞는 모든 축제 검색
    """
    print('searchAllFestival')
    url=f'http://api.visitkorea.or.kr/openapi/service/rest/KorService/searchFestival?numOfRows={numOfRows}&MobileOS=ETC&MobileApp=Festibot&serviceKey={serviceKey}&listYN={listYN}&arrange={arrange}&areaCode=1&eventStartDate={startDate}&eventEndDate={endDate}&pageNo=1&_type=json'
    resp = requests.get(url)
    data = resp.json()
    festivalInfo = data['response']['body']['items']['item']
    festivalList = []

    if content == False:
        count = int(data['response']['body']['totalCount'])
        if count > numOfRows:
            print('searchAllFestival 1 : ',festivalInfo)
            return festivalInfo
        else:
            for infoDec in festivalInfo:
                fixList = { 
                    'cat3':infoDec['cat3'], 'firstimage':infoDec['firstimage'], 'title':infoDec['title'],
                    'eventenddate':infoDec['eventenddate'], 'eventstartdate':infoDec['eventstartdate'],
                    'addr1':infoDec['addr1'] } #, 'tel':infoDec['tel'] }
                festivalList.append(copy.deepcopy(fixList))
            print('searchAllFestival 2 : ',festivalList)
            return festivalList
    else:
        rootCount = int(data['response']['body']['totalCount'])//20
        for i in range(1, rootCount+1):
            festivalList.extend(copy.deepcopy(searchContentFestival(user_id, startDate, endDate, content, i)))
        print('searchAllFestival 3 : ',festivalList)
        return festivalList

def showFestivalList(user_id, festivalList):
    i = 1
    resultShow=''
    for val in festivalList:
        resultShow = resultShow + f'{i}. ' + val['title'] + '\n'
        i += 1
    send_message(user_id, resultShow)

    print('festival_list_date 1 : ',festivalList)
    return festivalList

# 축제 갯수를 이용하여 조건 추가 여부 판단
def festival_list_date(user_id, **kwargs):
    """
    축제 갯수를 이용하여 조건 추가 여부 판단
    """
    print('festival_list_date')
    startDate = datetime.today() - timedelta(days=datetime.today().weekday())
    endDate = datetime.today() - timedelta(days=datetime.today().weekday()-7)
    startDate = startDate.strftime("%Y%m%d")
    endDate = endDate.strftime("%Y%m%d")

    if (not stateDB.isnull().loc[user_id, 'eventStartDate']) and (not stateDB.isnull().loc[user_id, 'contentCode']):
        festivalList = searchAllFestival(user_id, stateDB.loc[user_id, 'eventStartDate'], stateDB.loc[user_id, 'eventEndDate'], stateDB.loc[user_id, 'contentCode'])
        # 축제 갯수가 20개보다 많을 때
        if len(festivalList) > numOfRows:
            send_message(user_id, '일정과 축제 종류까지 선택했는데도 축제가 너무 많네요. 하지만 괜찮아요 가장 인기있는 축제 20개를 알려드릴께요 ! 이중에는 재미있는 축제가 너무너무 많답니다.')
            festivalList = showFestivalList(user_id, festivalList)
        
            print('festival_list_date 1 : ',festivalList)
            return festivalList
        else:
            festivalList = showFestivalList(user_id, festivalList)
            
            print('festival_list_date 2 : ',festivalList)
            return festivalList
    elif stateDB.isnull().loc[user_id, 'contentCode']:
        festivalList = searchAllFestival(user_id, stateDB.loc[user_id, 'eventStartDate'], stateDB.loc[user_id, 'eventEndDate'], False)
        # 축제 갯수가 20개보다 많을 때
        if len(festivalList) > numOfRows:
            send_message(user_id, '앗! 검색 결과가 너무 많아요. 다른 조건도 입력해 주세요. 축제 하면 전통행사 아니겠어요? 조건에 전통행사를 넣어보는 것도 추천드려요.')    
            
            print('festival_list_date 0 : ',festivalList)
            return '0'
        else:
            festivalList = showFestivalList(user_id, festivalList)
            
            print('festival_list_date 3 : ',festivalList)
            return festivalList
    elif stateDB.isnull().loc[user_id, 'eventStartDate']:
        festivalList = searchAllFestival(user_id, startDate, endDate, stateDB.loc[user_id, 'contentCode'])
        # 축제 갯수가 20개보다 많을 때
        if len(festivalList) > numOfRows:
            send_message(user_id, '앗! 검색 결과가 너무 많아요. 다른 조건도 입력해 주세요. 조건에 한 달 이내를 넣는 건 어떠세요? 이번 달에 재미있는 축제가 많아요!')
            print('festival_list_date 0 : ',festivalList)
            return '0'
        else:
            festivalList = showFestivalList(user_id, festivalList)
            
            print('festival_list_date 4 : ',festivalList)
            return festivalList
        
# 사용자가 선택한 축제의 상세 정보 출력
def choice_detailFestival(user_id, festivalList, msg):
    """
    사용자가 선택한 축제의 상세 정보 출력
    """
    print('choice_detailFestival')
    print('choice_detailFestival : ', festivalList)
    if len(festivalList) > int(msg):
        detailFestival = festivalList[int(msg)]
        url = 'https://api.telegram.org/bot{token}/sendPhoto'.format(token=API_KEY)
        params = {'chat_id': user_id, 'photo':detailFestival["firstimage"]}
        requests.post(url, json=params)
        url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
        msg = f'축제 종류 : {detailFestival["cat3"]}\n축제 이름 : {detailFestival["title"]}\n축제 기간 : {detailFestival["eventstartdate"]} ~ {detailFestival["eventenddate"]}\n주소 : {detailFestival["addr1"]}' #\nTel : {detailFestival["tel"]}'
        params = {'chat_id': user_id, 'text':msg}
        requests.post(url, json=params)
        send_message(user_id, '소개해드린 축제에 가고싶으신가요??')
        stateDB.loc[user_id] = np.nan
    else:
        send_message(user_id, '올바른 축제 번호를 입력해주세요 !')

# 가고 싶은 축제 선택
def choice_likeFestival(user_id):
    """
    가고 싶은 축제 선택
    """
    print('choice_likeFestival')
    for row in userInfoDB.rows:  
        if row[0].value == user_id: # 방문했던 적이 있었던 사람, 리스트로 저장되므로 리스트로 접근
            userRow = row[0].row()
    
    index = contentListCode.index(stateDB.loc[user_id, 'contentCode'])
    userInfoDB.cell(row=userRow, column=index+3).value += 1
    db.save(EXCEL_FILE_NAME)
    return False


############################## Main ##############################

# 사용자의 버튼 클릭을 바탕으로 상태코드를 부여.
# D : 기간 입력 필요/ C : 종류 입력 필요
# user_id = 사용자 아이디 코드, button_call : 버튼 입력 내역
def set_stateCode_button(user_id, msg, stateCode):
    """
    사용자의 버튼 클릭을 바탕으로 상태코드를 부여.
    D : 기간 입력 필요/ C : 종류 입력 필요
    user_id = 사용자 아이디 코드, button_call : 버튼 입력 내역
    """
    print('set_stateCode_button')
    
    # 최초 축제 기간 또는 종류 검색 선택
    # 축제 기간 선택
    if stateCode == 'D':
        send_message(user_id,'어떤 날짜에 놀러가고 싶어요? 아래 버튼으로 정할 수 있구, 특정 기간을 정하고 싶으면 YYYYMMDD-YYMMDD로 입력해줘요 !')
        choice_calendarDate(user_id, msg)
    # 축제 종류 선택
    elif stateCode == 'C':
        send_message(user_id, '어떤 축제에 놀러가고 싶어요?\n제가 축제 종류를 알려드릴께요 !')
        send_message(user_id, '1. 문화관광\n2. 일반\n3. 전통공연\n4. 연극\n5. 뮤지컬\n6. 오페라\n7. 전시회\n8. 박람회\n9. 컨벤션\n10. 무용\n11. 클래식음악회\n12. 대중콘서트\n13. 영화\n14. 스포츠경기\n15. 기타행사')
        send_message(user_id, '번호 또는 축제 종류를 적어주세요 !')
        choice_contentCode(user_id, msg)


# 경로 설정, URL 설정
@app_festibot.route('/', methods=['POST', 'GET'])
def index():
    if request.method == 'POST':
        message = request.get_json() # 사용자 대화 내용 불러옴
        user_id, userName, msg = parse_message(message)
            
        # 유저 정보를 확인하여, 기존 유저 또는 첫 방문자인지 검사
        thisUserIsFirst(user_id, userName)
        
        if msg == '종료':
            print(stateDB)
            stateDB.drop([user_id])
            print(stateDB)
            send_message(user_id, '종료할께요 ! 다음에 봬요 ~')

        print(stateDB)
        
        # 상태코드 확인
        if user_id in stateDB.index:
            # 상태 코드가 비어있을 때
            if stateDB.isnull().loc[user_id, 'stateCode']:
                # 일자 선택을 완료함
                if (not stateDB.isnull().loc[user_id, 'eventStartDate']) and stateDB.isnull().loc[user_id, 'contentCode']:
                    flag = festival_list_date(user_id, eventStartDate=stateDB.loc[user_id, 'eventStartDate'], eventEndDate=stateDB.loc[user_id, 'eventEndDate'])
                    print('flag :', flag)
                    if flag == '0':
                        stateDB.loc[user_id, 'stateCode'] = 'C'
                    else:
                        choice_detailFestival(user_id, flag, msg)

                # 종류 선택을 완료함
                elif (not stateDB.isnull().loc[user_id, 'contentCode']) and stateDB.isnull().loc[user_id, 'eventStartDate']:
                    flag = festival_list_date(user_id, content=stateDB.loc[user_id, 'contentCode'])
                    print('flag :', flag)
                    if flag == '0':
                        stateDB.loc[user_id, 'stateCode'] = 'D'
                    else:
                        choice_detailFestival(user_id, flag, msg)

                elif (not stateDB.isnull().loc[user_id, 'eventStartDate']) and (not stateDB.isnull().loc[user_id, 'contentCode']):
                    flag = festival_list_date(user_id, eventStartDate=stateDB.loc[user_id, 'eventStartDate'], eventEndDate=stateDB.loc[user_id, 'eventEndDate'], content=stateDB.loc[user_id, 'contentCode'])
                    print('flag :', flag)
                    choice_detailFestival(user_id, flag, msg)

                else:
                    # 축제 기간 또는 종류 검색 선택 시, 코드 부여
                    if msg == '축제 기간':
                        stateDB.loc[user_id, 'stateCode'] = 'D'
                    elif msg == '축제 종류':
                        stateDB.loc[user_id, 'stateCode'] = 'C'
                    elif msg == 'Yes':
                        stateDB.loc[user_id, 'stateCode'] = 'L'
                    else:
                        click_buttonFirst(user_id, msg)

            # 상태 코드가 비어있지 않을 때
            # 축제 기간 선택 필요
            if stateDB.loc[user_id, 'stateCode'] == 'D':
                if msg in dateList or dateWrite.match(msg):
                    choice_fixCalendarDate(user_id, msg)
                else:
                    set_stateCode_button(user_id, msg, 'D')
                
            # 축제 종류 선택 필요
            elif stateDB.loc[user_id, 'stateCode'] == 'C':
                set_stateCode_button(user_id, msg, 'C')

            elif stateDB.loc[user_id, 'stateCode'] == 'L':
                choice_likeFestival(user_id)
                stateDB.loc[user_id] = np.nan

            else:
                send_message(user_id, '다시 입력해주세요')
        
        return Response('ok', status=200)
    else:
        return 'Hello World!'

if __name__ == '__main__':
    app_festibot.run(port=5000)
